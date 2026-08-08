from __future__ import annotations

import json
import os
import shutil
import subprocess
from importlib import metadata
from pathlib import Path

import pytest

BIDS_1_11_1_SCHEMA_URL = (
    "https://raw.githubusercontent.com/bids-standard/bids-schema/"
    "34d59276aa8f34d3e3b2f17723183b5c7ecc1efb/versions/1.11.1/schema.json"
)


def _stat_fingerprint(path: Path) -> tuple[int, int]:
    stat = path.stat()
    return stat.st_size, stat.st_mtime_ns


def _write_synthetic_bids(root: Path) -> Path:
    root.mkdir()
    (root / "dataset_description.json").write_text(
        json.dumps({"Name": "Synthetic tool test", "BIDSVersion": "1.11.1"}),
        encoding="utf-8",
    )
    (root / "README").write_text("Synthetic metadata-only test.\n", encoding="utf-8")
    (root / "task-test_eeg.json").write_text(
        json.dumps(
            {
                "TaskName": "test",
                "SamplingFrequency": 256,
                "PowerLineFrequency": 50,
                "EEGReference": "Cz",
                "SoftwareFilters": "n/a",
            }
        ),
        encoding="utf-8",
    )
    eeg = root / "sub-001" / "eeg"
    eeg.mkdir(parents=True)
    selected = eeg / "sub-001_task-test_run-1_eeg.edf"
    selected.write_bytes(b"container payload is not opened by this PyBIDS test")
    selected.with_suffix(".json").write_text(
        json.dumps({"SamplingFrequency": 512}), encoding="utf-8"
    )
    (eeg / "sub-001_task-test_run-1_channels.tsv").write_text(
        "name\ttype\tunits\tstatus\nFz\tEEG\tuV\tgood\nCz\tEEG\tuV\tgood\n",
        encoding="utf-8",
    )
    return selected


def test_pybids_resolves_inheritance_without_opening_payload(tmp_path: Path) -> None:
    from bids import BIDSLayout

    root = tmp_path / "ds-test"
    selected = _write_synthetic_bids(root)
    before = _stat_fingerprint(selected)
    layout = BIDSLayout(
        root,
        validate=False,
        config="bids-schema",
        derivatives=False,
    )

    files = layout.get(
        subject="001",
        task="test",
        run=1,
        suffix="eeg",
        extension="edf",
        return_type="filename",
    )
    metadata = layout.get_metadata(files[0])

    assert _stat_fingerprint(selected) == before
    assert metadata["SamplingFrequency"] == 512
    assert metadata["EEGReference"] == "Cz"
    assert metadata["SoftwareFilters"] == "n/a"
    assert layout.parse_file_entities(files[0])["subject"] == "001"


def test_mne_hdf5_extra_reads_matlab_v73_without_matlab(
    tmp_path: Path,
) -> None:
    import h5py
    import numpy as np
    from pymatreader import read_mat

    mat_path = tmp_path / "synthetic_v73.mat"
    with h5py.File(mat_path, "w", userblock_size=512) as handle:
        signal = handle.create_dataset(
            "signal",
            data=np.arange(12, dtype=np.float64).reshape(4, 3),
        )
        signal.attrs["MATLAB_class"] = np.bytes_("double")
        sfreq = handle.create_dataset(
            "sfreq",
            data=np.array([[256.0]], dtype=np.float64),
        )
        sfreq.attrs["MATLAB_class"] = np.bytes_("double")

    description = (
        b"MATLAB 7.3 MAT-file, Platform: eeg-provenance test, "
        b"HDF5 schema 1.00 ."
    )
    header = (
        description.ljust(116, b" ")
        + bytes(8)
        + bytes([0, 2])
        + b"IM"
    )
    with mat_path.open("r+b") as handle:
        handle.write(header)

    assert h5py.is_hdf5(mat_path)
    result = read_mat(mat_path, variable_names=["signal", "sfreq"])

    assert metadata.version("mne") == "1.12.1"
    assert metadata.version("h5io")
    assert metadata.version("pymatreader")
    assert result["signal"].shape == (3, 4)
    assert float(np.asarray(result["sfreq"]).squeeze()) == 256.0


def test_official_bids_validator_binary_returns_json(tmp_path: Path) -> None:
    executable = shutil.which("bids-validator-deno")
    if executable is None:
        pytest.skip("official BIDS Validator binary is not installed")
    root = tmp_path / "ds-validator"
    root.mkdir()
    (root / "dataset_description.json").write_text(
        json.dumps({"Name": "Validator smoke test", "BIDSVersion": "1.11.1"}),
        encoding="utf-8",
    )
    (root / "README").write_text("No recordings in smoke fixture.\n", encoding="utf-8")

    completed = subprocess.run(
        [
            executable,
            str(root),
            "--format",
            "json",
            "--max-rows",
            "0",
            "--ignoreNiftiHeaders",
        ],
        check=False,
        capture_output=True,
        text=True,
        timeout=60,
    )

    assert completed.returncode in {0, 1}
    assert completed.stdout, completed.stderr
    result = json.loads(completed.stdout)
    assert {"issues", "summary"} <= result.keys()


@pytest.mark.network
def test_official_bids_validator_accepts_pinned_1_11_1_schema(
    tmp_path: Path,
) -> None:
    if os.environ.get("EEG_PROVENANCE_NETWORK_TESTS") != "1":
        pytest.skip("set EEG_PROVENANCE_NETWORK_TESTS=1 for live schema retrieval")
    executable = shutil.which("bids-validator-deno")
    if executable is None:
        pytest.skip("official BIDS Validator binary is not installed")
    root = tmp_path / "ds-validator-schema"
    root.mkdir()
    (root / "dataset_description.json").write_text(
        json.dumps({"Name": "Schema smoke test", "BIDSVersion": "1.11.1"}),
        encoding="utf-8",
    )
    (root / "README").write_text("Schema smoke fixture.\n", encoding="utf-8")

    completed = subprocess.run(
        [
            executable,
            str(root),
            "--schema",
            BIDS_1_11_1_SCHEMA_URL,
            "--format",
            "json",
            "--max-rows",
            "0",
            "--ignoreNiftiHeaders",
        ],
        check=False,
        capture_output=True,
        text=True,
        timeout=60,
    )

    assert completed.returncode in {0, 1}
    assert completed.stdout, completed.stderr
    result = json.loads(completed.stdout)
    assert {"issues", "summary"} <= result.keys()


@pytest.mark.adapters
def test_moabb_bci_adapter_contract_without_download() -> None:
    pytest.importorskip("moabb")
    from moabb.datasets import BNCI2014_001

    dataset = BNCI2014_001()

    assert len(dataset.subject_list) == 9
    assert dataset.event_id == {
        "left_hand": 1,
        "right_hand": 2,
        "feet": 3,
        "tongue": 4,
    }
    assert dataset.interval == [2, 6]


@pytest.mark.archive
def test_openneuro_metadata_with_pybids_is_read_only() -> None:
    from bids import BIDSLayout

    root = Path("X:/openneurodatasets/ds003061")
    description = root / "dataset_description.json"
    if not description.is_file():
        pytest.skip("OpenNeuro archive is not mounted")
    before = _stat_fingerprint(description)

    layout = BIDSLayout(
        root,
        validate=False,
        config="bids-schema",
        derivatives=False,
    )
    dataset = layout.get_dataset_description()
    sidecars = layout.get(
        subject="001",
        task="P300",
        run=1,
        suffix="eeg",
        extension="json",
        return_type="filename",
    )

    assert _stat_fingerprint(description) == before
    assert "auditory oddball" in dataset["Name"].casefold()
    eeg_sidecars = [path for path in sidecars if path.endswith("_eeg.json")]
    assert len(eeg_sidecars) == 1
    sidecar = json.loads(Path(eeg_sidecars[0]).read_text(encoding="utf-8"))
    assert sidecar["SamplingFrequency"] == 256


@pytest.mark.archive
def test_bci_iv_2a_with_mne_and_scipy_is_header_bounded() -> None:
    import mne
    from scipy.io import loadmat, whosmat

    root = Path("X:/data/bci_competition_iv/BCICIV_2a_gdf")
    recording = root / "A01T.gdf"
    labels = root / "true_labels_official" / "A01T.mat"
    if not recording.is_file() or not labels.is_file():
        pytest.skip("BCI Competition IV 2a archive is not mounted")
    before = (_stat_fingerprint(recording), _stat_fingerprint(labels))

    raw = mne.io.read_raw_gdf(recording, preload=False, verbose="ERROR")
    annotation_counts = {
        value: list(raw.annotations.description).count(value)
        for value in set(raw.annotations.description)
    }
    variables = whosmat(labels)
    class_labels = loadmat(labels, variable_names=["classlabel"])[
        "classlabel"
    ].reshape(-1)

    assert (_stat_fingerprint(recording), _stat_fingerprint(labels)) == before
    assert raw.preload is False
    assert raw.info["sfreq"] == 250
    assert len(raw.ch_names) == 25
    assert annotation_counts["768"] == 288
    assert variables == [("classlabel", (288, 1), "double")]
    assert {
        int(value): int((class_labels == value).sum())
        for value in sorted(set(class_labels))
    } == {1: 72, 2: 72, 3: 72, 4: 72}


@pytest.mark.archive
def test_seed_with_scipy_and_openpyxl_avoids_signal_arrays() -> None:
    from openpyxl import load_workbook
    from scipy.io import whosmat

    root = Path("X:/data/sjtu_emotion_eeg/SEED")
    recording = root / "Preprocessed_EEG" / "1_20131027.mat"
    channel_order = root / "channel-order.xlsx"
    stimuli = root / "seed-stimulation.xlsx"
    if not recording.is_file() or not channel_order.is_file() or not stimuli.is_file():
        pytest.skip("SEED archive is not mounted")
    before = (
        _stat_fingerprint(recording),
        _stat_fingerprint(channel_order),
        _stat_fingerprint(stimuli),
    )

    variables = whosmat(recording)
    workbook = load_workbook(channel_order, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        labels = [
            str(value).strip()
            for row in sheet.iter_rows(values_only=True)
            for value in row
            if value is not None and str(value).strip()
        ]
    finally:
        workbook.close()

    workbook = load_workbook(stimuli, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(min_row=1, max_row=18, values_only=True))
        stimulus_codes = [int(row[1]) for row in rows[1:16]]
        encoding_note = str(rows[17][0])
    finally:
        workbook.close()

    assert (
        _stat_fingerprint(recording),
        _stat_fingerprint(channel_order),
        _stat_fingerprint(stimuli),
    ) == before
    trial_arrays = [item for item in variables if "eeg" in item[0].casefold()]
    assert len(trial_arrays) == 15
    assert all(item[1][0] == 62 for item in trial_arrays)
    assert len(labels) == 62
    assert set(stimulus_codes) == {0, 1, 2}
    assert "0: sad, 1: neutral, 2: happy" in encoding_note


@pytest.mark.archive
@pytest.mark.parametrize(
    ("root", "relative"),
    [
        (
            Path("X:/data/tuh_eeg/tuh_eeg/v2.0.1"),
            Path("edf/000/aaaaaaaa/s001_2015/01_tcp_ar/aaaaaaaa_s001_t000.edf"),
        ),
        (
            Path("X:/data/tuh_eeg/tuh_eeg_abnormal/v3.0.1"),
            Path("edf/eval/normal/01_tcp_ar/aaaaaayx_s002_t000.edf"),
        ),
    ],
)
def test_tuh_edf_headers_with_pyedflib_are_per_signal_and_read_only(
    root: Path, relative: Path
) -> None:
    import pyedflib

    recording = root / relative
    if not recording.is_file():
        pytest.skip("TUH archive is not mounted")
    before = _stat_fingerprint(recording)

    reader = pyedflib.EdfReader(str(recording))
    try:
        headers = reader.getSignalHeaders()
    finally:
        reader.close()

    assert _stat_fingerprint(recording) == before
    assert headers
    assert any(header["label"].startswith("EEG") for header in headers)
    assert all(header["dimension"] for header in headers)
    assert all(header["sample_frequency"] > 0 for header in headers)


@pytest.mark.archive
def test_tuh_release_readmes_preserve_scale_and_split_contracts() -> None:
    tueg = Path("X:/data/tuh_eeg/tuh_eeg/v2.0.1/AAREADME.txt")
    tuab = Path("X:/data/tuh_eeg/tuh_eeg_abnormal/v3.0.1/AAREADME.txt")
    if not tueg.is_file() or not tuab.is_file():
        pytest.skip("TUH release metadata are not mounted")
    before = (_stat_fingerprint(tueg), _stat_fingerprint(tuab))

    tueg_text = tueg.read_text(encoding="utf-8", errors="replace")
    tuab_text = tuab.read_text(encoding="utf-8", errors="replace")

    assert (_stat_fingerprint(tueg), _stat_fingerprint(tuab)) == before
    assert "Number of Edf files: 69,670" in tueg_text
    assert "69672 out of 69672" in tueg_text
    assert "1,643 Gbytes" in tueg_text
    assert "100% disjoint" in tuab_text
    assert "54" in tuab_text and "both normal and abnormal" in tuab_text
